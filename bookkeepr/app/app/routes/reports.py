"""Report Generation Routes - Live Financial Dashboard"""
from flask import Blueprint, request, send_file, jsonify, render_template
from flask_login import login_required, current_user
from app.models import Transaction, Account, Company
from extensions import db
from datetime import datetime, timedelta
import io
import csv

bp = Blueprint('reports', __name__)


# ============================================================================
# HTML DASHBOARD ROUTES
# ============================================================================

@bp.route('/dashboard/reports')
@login_required
def reports_dashboard():
    """Interactive reports dashboard page"""
    return render_template('dashboard/reports.html')


# ============================================================================
# API ENDPOINTS - LIVE FINANCIAL REPORTS
# ============================================================================

@bp.route('/api/v1/reports/pnl')
@login_required
def api_pnl():
    """Get Profit & Loss data with date filtering"""
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    data = generate_pnl_data(start_date, end_date)
    return jsonify(data)


@bp.route('/api/v1/reports/balance')
@login_required
def api_balance():
    """Get Balance Sheet data as of a specific date"""
    as_of = request.args.get('as_of')
    data = generate_balance_data(as_of)
    return jsonify(data)


@bp.route('/api/v1/reports/cashflow')
@login_required
def api_cashflow():
    """Get Cash Flow data with period grouping"""
    period = request.args.get('period', 'monthly')  # monthly or weekly
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    data = generate_cashflow_data(period, start_date, end_date)
    return jsonify(data)


@bp.route('/api/v1/reports/trial-balance')
@login_required
def api_trial_balance():
    """Get Trial Balance data"""
    as_of = request.args.get('as_of')
    data = generate_trial_balance(as_of)
    return jsonify(data)


@bp.route('/api/v1/reports/ledger')
@login_required
def api_ledger():
    """Get General Ledger data with filtering"""
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    category = request.args.get('category')
    account = request.args.get('account')
    search = request.args.get('search')
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 50, type=int)
    
    data = generate_ledger_data(start_date, end_date, category, account, search, page, per_page)
    return jsonify(data)


# ============================================================================
# DOWNLOAD ENDPOINTS
# ============================================================================

@bp.route('/api/v1/reports/download/<report_type>')
@login_required
def download_report(report_type):
    """Generate downloadable report"""
    format_type = request.args.get('format', 'csv')  # csv, pdf, excel
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    if report_type == 'pnl':
        data = generate_pnl_data(start_date, end_date)
        filename = f'pnl_report_{datetime.now().strftime("%Y%m%d")}'
    elif report_type == 'balance':
        as_of = request.args.get('as_of')
        data = generate_balance_data(as_of)
        filename = f'balance_sheet_{datetime.now().strftime("%Y%m%d")}'
    elif report_type == 'cashflow':
        period = request.args.get('period', 'monthly')
        data = generate_cashflow_data(period, start_date, end_date)
        filename = f'cashflow_{datetime.now().strftime("%Y%m%d")}'
    elif report_type == 'trial-balance':
        as_of = request.args.get('as_of')
        data = generate_trial_balance(as_of)
        filename = f'trial_balance_{datetime.now().strftime("%Y%m%d")}'
    elif report_type == 'ledger':
        data = generate_ledger_data(start_date, end_date)
        filename = f'general_ledger_{datetime.now().strftime("%Y%m%d")}'
    else:
        return jsonify({'error': 'Invalid report type'}), 400
    
    if format_type == 'csv':
        return create_csv_download(data, filename + '.csv')
    elif format_type == 'excel':
        return create_excel_download(data, filename + '.xlsx')
    elif format_type == 'pdf':
        return create_pdf_download(data, filename + '.pdf')
    else:
        return jsonify({'error': 'Invalid format'}), 400


# ============================================================================
# REPORT DATA GENERATORS
# ============================================================================

def get_company():
    """Get user's connected company (or any active one as fallback)"""
    company = Company.query.filter_by(
        user_id=current_user.id,
        is_active=True,
        is_connected=True,
    ).first()
    if not company:
        company = Company.query.filter_by(user_id=current_user.id, is_active=True).first()
    if not company:
        company = Company.query.filter_by(user_id=current_user.id).first()
    return company


def get_date_range(start_date_str=None, end_date_str=None, default_days=365):
    """Parse date range with defaults"""
    if end_date_str:
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
    else:
        end_date = datetime.now()
    
    if start_date_str:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
    else:
        start_date = end_date - timedelta(days=default_days)
    
    return start_date.date(), end_date.date()


def generate_pnl_data(start_date_str=None, end_date_str=None):
    """Generate Profit & Loss data with date filtering"""
    company = get_company()
    if not company:
        return {
            'report_type': 'Profit & Loss',
            'period': 'No company found',
            'generated_at': datetime.now().isoformat(),
            'summary': {'total_income': 0, 'total_expenses': 0, 'net_profit': 0, 'profit_margin': 0},
            'income_accounts': [],
            'expense_accounts': []
        }
    
    start_date, end_date = get_date_range(start_date_str, end_date_str, 365)
    
    # Get all transactions in date range
    transactions = Transaction.query.filter(
        Transaction.company_id == company.id,
        Transaction.transaction_date >= start_date,
        Transaction.transaction_date <= end_date
    ).all()
    
    # Group by account/category
    income_by_account = {}
    expense_by_account = {}
    
    total_income = 0
    total_expenses = 0
    
    for t in transactions:
        amount = float(t.amount) if t.amount else 0
        account_name = t.account_name or 'Uncategorized'
        
        if amount > 0:
            total_income += amount
            income_by_account[account_name] = income_by_account.get(account_name, 0) + amount
        else:
            total_expenses += abs(amount)
            expense_by_account[account_name] = expense_by_account.get(account_name, 0) + abs(amount)
    
    net_profit = total_income - total_expenses
    profit_margin = (net_profit / total_income * 100) if total_income > 0 else 0
    
    # Sort by amount descending
    income_accounts = [
        {'name': name, 'amount': amount}
        for name, amount in sorted(income_by_account.items(), key=lambda x: x[1], reverse=True)
    ]
    expense_accounts = [
        {'name': name, 'amount': amount}
        for name, amount in sorted(expense_by_account.items(), key=lambda x: x[1], reverse=True)
    ]
    
    # Build line_items in unified format for the frontend
    line_items = []
    for it in income_accounts:
        line_items.append({'category': it['name'], 'amount': it['amount']})
    for it in expense_accounts:
        line_items.append({'category': it['name'], 'amount': -it['amount']})
    
    return {
        'report_type': 'Profit & Loss',
        'period': f'{start_date.isoformat()} to {end_date.isoformat()}',
        'generated_at': datetime.now().isoformat(),
        'totals': {
            'revenue': round(total_income, 2),
            'expenses': round(total_expenses, 2),
            'net_income': round(net_profit, 2),
            'profit_margin': round(profit_margin, 2),
            'transaction_count': len(transactions),
        },
        'line_items': line_items,
        'summary': {
            'total_income': round(total_income, 2),
            'total_expenses': round(total_expenses, 2),
            'net_profit': round(net_profit, 2),
            'profit_margin': round(profit_margin, 2)
        },
        'income_accounts': income_accounts,
        'expense_accounts': expense_accounts
    }


def generate_balance_data(as_of_str=None):
    """Generate Balance Sheet data as of a specific date"""
    company = get_company()
    if not company:
        return {
            'report_type': 'Balance Sheet',
            'as_of': 'No company found',
            'generated_at': datetime.now().isoformat(),
            'summary': {'total_assets': 0, 'total_liabilities': 0, 'total_equity': 0, 'balance_check': 0},
            'assets': [],
            'liabilities': [],
            'equity': []
        }
    
    if as_of_str:
        as_of = datetime.strptime(as_of_str, '%Y-%m-%d').date()
    else:
        as_of = datetime.now().date()
    
    # Get accounts with their current balances
    accounts = Account.query.filter_by(company_id=company.id).all()
    
    assets = []
    liabilities = []
    equity = []
    
    total_assets = 0
    total_liabilities = 0
    total_equity = 0
    
    for account in accounts:
        balance = float(account.current_balance) if account.current_balance else 0
        account_data = {
            'name': account.name,
            'account_type': account.account_type or 'Unknown',
            'balance': round(balance, 2)
        }
        
        classification = (account.classification or account.account_type or '').lower()
        
        if 'asset' in classification:
            assets.append(account_data)
            total_assets += balance
        elif 'liability' in classification:
            liabilities.append(account_data)
            total_liabilities += abs(balance)
        elif 'equity' in classification:
            equity.append(account_data)
            total_equity += balance
    
    # Sort by balance descending
    assets.sort(key=lambda x: x['balance'], reverse=True)
    liabilities.sort(key=lambda x: x['balance'], reverse=True)
    equity.sort(key=lambda x: x['balance'], reverse=True)
    
    balance_check = total_assets - (total_liabilities + total_equity)
    
    return {
        'report_type': 'Balance Sheet',
        'as_of': as_of.isoformat(),
        'generated_at': datetime.now().isoformat(),
        'summary': {
            'total_assets': round(total_assets, 2),
            'total_liabilities': round(total_liabilities, 2),
            'total_equity': round(total_equity, 2),
            'balance_check': round(balance_check, 2)
        },
        'assets': assets,
        'liabilities': liabilities,
        'equity': equity
    }


def generate_cashflow_data(period='monthly', start_date_str=None, end_date_str=None):
    """Generate Cash Flow data with period grouping"""
    company = get_company()
    if not company:
        return {
            'report_type': 'Cash Flow',
            'period': 'No company found',
            'generated_at': datetime.now().isoformat(),
            'summary': {'operating_cash_flow': 0, 'investing_cash_flow': 0, 'financing_cash_flow': 0, 'net_cash_flow': 0},
            'periods': []
        }
    
    start_date, end_date = get_date_range(start_date_str, end_date_str, 365)
    
    # Get transactions in date range
    transactions = Transaction.query.filter(
        Transaction.company_id == company.id,
        Transaction.transaction_date >= start_date,
        Transaction.transaction_date <= end_date
    ).all()
    
    # Group by period
    from collections import defaultdict
    periods = defaultdict(lambda: {'operating': 0, 'investing': 0, 'financing': 0, 'total': 0})
    
    operating_categories = ['Sales', 'Services', 'Operating Expenses', 'Cost of Goods Sold', 
                            'Supplies', 'Utilities', 'Rent', 'Salaries', 'Marketing']
    investing_categories = ['Equipment', 'Investments', 'Fixed Assets', 'Machinery', 'Vehicles']
    financing_categories = ['Loans', 'Equity', 'Dividends', 'Owner Contribution', 'Loan Repayment']
    
    total_operating = 0
    total_investing = 0
    total_financing = 0
    
    for t in transactions:
        amount = float(t.amount) if t.amount else 0
        category = (t.category or 'Uncategorized').lower()
        
        # Determine period key
        if period == 'weekly':
            # Get week start (Monday)
            days_since_monday = t.transaction_date.weekday()
            week_start = t.transaction_date - timedelta(days=days_since_monday)
            period_key = week_start.isoformat()
            period_label = f"Week of {week_start.strftime('%b %d, %Y')}"
        else:
            period_key = t.transaction_date.strftime('%Y-%m')
            period_label = t.transaction_date.strftime('%b %Y')
        
        # Classify
        if any(op_cat.lower() in category for op_cat in operating_categories):
            periods[period_key]['operating'] += amount
            total_operating += amount
        elif any(inv_cat.lower() in category for inv_cat in investing_categories):
            periods[period_key]['investing'] += amount
            total_investing += amount
        elif any(fin_cat.lower() in category for fin_cat in financing_categories):
            periods[period_key]['financing'] += amount
            total_financing += amount
        else:
            # Default to operating
            periods[period_key]['operating'] += amount
            total_operating += amount
        
        periods[period_key]['total'] += amount
        periods[period_key]['label'] = period_label
    
    # Convert to list sorted by period
    period_list = [
        {
            'period': key,
            'label': data['label'],
            'operating_cash_flow': round(data['operating'], 2),
            'investing_cash_flow': round(data['investing'], 2),
            'financing_cash_flow': round(data['financing'], 2),
            'net_cash_flow': round(data['total'], 2)
        }
        for key, data in sorted(periods.items())
    ]
    
    return {
        'report_type': 'Cash Flow',
        'period_type': period,
        'period': f'{start_date.isoformat()} to {end_date.isoformat()}',
        'generated_at': datetime.now().isoformat(),
        'summary': {
            'operating_cash_flow': round(total_operating, 2),
            'investing_cash_flow': round(total_investing, 2),
            'financing_cash_flow': round(total_financing, 2),
            'net_cash_flow': round(total_operating + total_investing + total_financing, 2)
        },
        'periods': period_list
    }


def generate_trial_balance(as_of_str=None):
    """Generate Trial Balance"""
    company = get_company()
    if not company:
        return {
            'report_type': 'Trial Balance',
            'as_of': 'No company found',
            'generated_at': datetime.now().isoformat(),
            'summary': {'total_debits': 0, 'total_credits': 0, 'difference': 0},
            'accounts': []
        }
    
    if as_of_str:
        as_of = datetime.strptime(as_of_str, '%Y-%m-%d').date()
    else:
        as_of = datetime.now().date()
    
    accounts = Account.query.filter_by(company_id=company.id).all()
    
    account_list = []
    total_debits = 0
    total_credits = 0
    
    for account in accounts:
        balance = float(account.current_balance) if account.current_balance else 0
        
        if balance > 0:
            debit = balance
            credit = 0
            total_debits += balance
        else:
            debit = 0
            credit = abs(balance)
            total_credits += abs(balance)
        
        account_list.append({
            'name': account.name,
            'account_type': account.account_type or 'Unknown',
            'account_number': account.account_number or '',
            'debit': round(debit, 2),
            'credit': round(credit, 2),
            'balance': round(balance, 2)
        })
    
    # Sort by account type then name
    account_list.sort(key=lambda x: (x['account_type'], x['name']))
    
    return {
        'report_type': 'Trial Balance',
        'as_of': as_of.isoformat(),
        'generated_at': datetime.now().isoformat(),
        'summary': {
            'total_debits': round(total_debits, 2),
            'total_credits': round(total_credits, 2),
            'difference': round(total_debits - total_credits, 2)
        },
        'accounts': account_list
    }


def generate_ledger_data(start_date_str=None, end_date_str=None, category=None, 
                         account=None, search=None, page=1, per_page=50):
    """Generate General Ledger with filtering and pagination"""
    company = get_company()
    if not company:
        return {
            'report_type': 'General Ledger',
            'period': 'No company found',
            'generated_at': datetime.now().isoformat(),
            'transactions': [],
            'total': 0,
            'pages': 0,
            'current_page': 1
        }
    
    # Build query
    query = Transaction.query.filter_by(company_id=company.id)
    
    if start_date_str:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        query = query.filter(Transaction.transaction_date >= start_date)
    
    if end_date_str:
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        query = query.filter(Transaction.transaction_date <= end_date)
    
    if category:
        query = query.filter(Transaction.category.ilike(f'%{category}%'))
    
    if account:
        query = query.filter(Transaction.account_name.ilike(f'%{account}%'))
    
    if search:
        query = query.filter(
            db.or_(
                Transaction.description.ilike(f'%{search}%'),
                Transaction.vendor_name.ilike(f'%{search}%'),
                Transaction.memo.ilike(f'%{search}%')
            )
        )
    
    # Get total count before pagination
    total = query.count()
    
    # Apply pagination
    transactions = query.order_by(Transaction.transaction_date.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )
    
    return {
        'report_type': 'General Ledger',
        'generated_at': datetime.now().isoformat(),
        'transactions': [{
            'id': t.id,
            'date': t.transaction_date.isoformat() if t.transaction_date else None,
            'description': t.description,
            'vendor_name': t.vendor_name,
            'amount': float(t.amount) if t.amount else 0,
            'category': t.category or 'Uncategorized',
            'account': t.account_name or 'Uncategorized',
            'transaction_type': t.transaction_type,
            'categorization_status': t.categorization_status,
            'review_status': t.review_status
        } for t in transactions.items],
        'total': total,
        'pages': transactions.pages,
        'current_page': page,
        'per_page': per_page
    }


# ============================================================================
# EXPORT FORMATTERS
# ============================================================================

def create_csv_download(data, filename):
    """Create CSV file for download"""
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Write metadata
    writer.writerow(['Report Type:', data.get('report_type', 'Report')])
    if 'period' in data:
        writer.writerow(['Period:', data['period']])
    if 'as_of' in data:
        writer.writerow(['As Of:', data['as_of']])
    writer.writerow(['Generated:', data.get('generated_at', datetime.now().isoformat())])
    writer.writerow([])
    
    # Write summary
    if 'summary' in data:
        writer.writerow(['Summary'])
        for key, value in data['summary'].items():
            writer.writerow([key.replace('_', ' ').title(), value])
        writer.writerow([])
    
    # Write accounts/transactions based on report type
    report_type = data.get('report_type', '').lower()
    
    if 'profit & loss' in report_type or 'pnl' in report_type:
        if data.get('income_accounts'):
            writer.writerow(['Income Accounts'])
            writer.writerow(['Account', 'Amount'])
            for acc in data['income_accounts']:
                writer.writerow([acc['name'], acc['amount']])
            writer.writerow([])
        if data.get('expense_accounts'):
            writer.writerow(['Expense Accounts'])
            writer.writerow(['Account', 'Amount'])
            for acc in data['expense_accounts']:
                writer.writerow([acc['name'], acc['amount']])
    
    elif 'balance sheet' in report_type:
        for section in ['assets', 'liabilities', 'equity']:
            if data.get(section):
                writer.writerow([section.title()])
                writer.writerow(['Account', 'Type', 'Balance'])
                for acc in data[section]:
                    writer.writerow([acc['name'], acc['account_type'], acc['balance']])
                writer.writerow([])
    
    elif 'trial balance' in report_type:
        if data.get('accounts'):
            writer.writerow(['Accounts'])
            writer.writerow(['Account', 'Account Type', 'Account #', 'Debit', 'Credit', 'Balance'])
            for acc in data['accounts']:
                writer.writerow([
                    acc['name'], acc['account_type'], acc['account_number'],
                    acc['debit'], acc['credit'], acc['balance']
                ])
    
    elif 'cash flow' in report_type:
        if data.get('periods'):
            writer.writerow(['Cash Flow by Period'])
            writer.writerow(['Period', 'Operating', 'Investing', 'Financing', 'Net'])
            for period in data['periods']:
                writer.writerow([
                    period['label'],
                    period['operating_cash_flow'],
                    period['investing_cash_flow'],
                    period['financing_cash_flow'],
                    period['net_cash_flow']
                ])
    
    elif 'ledger' in report_type or 'transactions' in data:
        if data.get('transactions'):
            writer.writerow(['Transactions'])
            writer.writerow(['Date', 'Description', 'Vendor', 'Amount', 'Category', 'Account', 'Type', 'Status'])
            for t in data['transactions']:
                writer.writerow([
                    t.get('date', ''),
                    t.get('description', ''),
                    t.get('vendor_name', ''),
                    t.get('amount', 0),
                    t.get('category', ''),
                    t.get('account', ''),
                    t.get('transaction_type', ''),
                    t.get('review_status', '')
                ])
    
    output.seek(0)
    return send_file(
        io.BytesIO(output.getvalue().encode('utf-8-sig')),
        mimetype='text/csv; charset=utf-8',
        as_attachment=True,
        download_name=filename
    )


def create_excel_download(data, filename):
    """Create Excel file for download"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        # Fallback to CSV if openpyxl not available
        return create_csv_download(data, filename.replace('.xlsx', '.csv'))
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"
    
    # Header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    row = 1
    
    # Title
    ws.merge_cells('A1:D1')
    ws['A1'] = data.get('report_type', 'Report')
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center')
    row = 3
    
    # Metadata
    ws[f'A{row}'] = 'Period:'
    ws[f'B{row}'] = data.get('period', data.get('as_of', 'N/A'))
    row += 1
    ws[f'A{row}'] = 'Generated:'
    ws[f'B{row}'] = data.get('generated_at', datetime.now().isoformat())
    row += 2
    
    # Summary
    if 'summary' in data:
        ws[f'A{row}'] = 'Summary'
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        for key, value in data['summary'].items():
            ws[f'A{row}'] = key.replace('_', ' ').title()
            ws[f'B{row}'] = value
            row += 1
        row += 1
    
    # Data based on report type
    report_type = data.get('report_type', '').lower()
    
    if data.get('transactions'):
        headers = ['Date', 'Description', 'Vendor', 'Amount', 'Category', 'Account', 'Type', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
        row += 1
        
        for t in data['transactions']:
            ws.cell(row=row, column=1, value=t.get('date', ''))
            ws.cell(row=row, column=2, value=t.get('description', ''))
            ws.cell(row=row, column=3, value=t.get('vendor_name', ''))
            ws.cell(row=row, column=4, value=t.get('amount', 0))
            ws.cell(row=row, column=5, value=t.get('category', ''))
            ws.cell(row=row, column=6, value=t.get('account', ''))
            ws.cell(row=row, column=7, value=t.get('transaction_type', ''))
            ws.cell(row=row, column=8, value=t.get('review_status', ''))
            row += 1
    
    elif data.get('accounts'):
        headers = ['Account', 'Type', 'Account #', 'Debit', 'Credit', 'Balance']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
        row += 1
        
        for acc in data['accounts']:
            ws.cell(row=row, column=1, value=acc.get('name', ''))
            ws.cell(row=row, column=2, value=acc.get('account_type', ''))
            ws.cell(row=row, column=3, value=acc.get('account_number', ''))
            ws.cell(row=row, column=4, value=acc.get('debit', 0))
            ws.cell(row=row, column=5, value=acc.get('credit', 0))
            ws.cell(row=row, column=6, value=acc.get('balance', 0))
            row += 1
    
    # Auto-adjust column widths
    for col in range(1, ws.max_column + 1):
        max_length = 0
        column = get_column_letter(col)
        for cell in ws[column]:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


def create_pdf_download(data, filename):
    """Create PDF file for download"""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
    except ImportError:
        # Fallback to CSV if reportlab not available
        return create_csv_download(data, filename.replace('.pdf', '.csv'))
    
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#2563EB'),
        spaceAfter=20,
        alignment=1  # Center
    )
    elements.append(Paragraph(data.get('report_type', 'Report'), title_style))
    elements.append(Spacer(1, 0.1*inch))
    
    # Metadata
    meta_style = ParagraphStyle(
        'MetaStyle',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.gray,
        spaceAfter=5
    )
    period = data.get('period', data.get('as_of', 'N/A'))
    elements.append(Paragraph(f"Period: {period}", meta_style))
    elements.append(Paragraph(f"Generated: {data.get('generated_at', datetime.now().isoformat())}", meta_style))
    elements.append(Spacer(1, 0.2*inch))
    
    # Summary
    if 'summary' in data:
        elements.append(Paragraph("Summary", styles['Heading2']))
        summary_data = [['Item', 'Value']]
        for key, value in data['summary'].items():
            summary_data.append([key.replace('_', ' ').title(), str(value)])
        
        summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563EB')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F3F4F6')),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#D1D5DB')),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('TOPPADDING', (0, 1), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 0.2*inch))
    
    # Transaction/Account data
    if data.get('transactions'):
        elements.append(Paragraph("Transactions", styles['Heading2']))
        table_data = [['Date', 'Description', 'Amount', 'Category', 'Account']]
        for t in data['transactions'][:100]:  # Limit to 100 for PDF
            table_data.append([
                t.get('date', '')[:10],
                t.get('description', '')[:40],
                f"${t.get('amount', 0):,.2f}",
                t.get('category', ''),
                t.get('account', '')
            ])
    elif data.get('accounts'):
        elements.append(Paragraph("Accounts", styles['Heading2']))
        table_data = [['Account', 'Type', 'Balance']]
        for acc in data['accounts'][:100]:
            table_data.append([
                acc.get('name', ''),
                acc.get('account_type', ''),
                f"${acc.get('balance', 0):,.2f}"
            ])
    else:
        table_data = []
    
    if table_data:
        table = Table(table_data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563EB')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F9FAFB')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E5E7EB')),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('TOPPADDING', (0, 1), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 3),
        ]))
        elements.append(table)
    
    doc.build(elements)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=filename
    )
